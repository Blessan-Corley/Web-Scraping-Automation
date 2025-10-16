"""
KIT Portal Student Data Automation - FULLY FIXED VERSION
Fixed: Login button click issue - now uses multiple click strategies
"""

import time
import os
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import re

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import easyocr
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import requests
from io import BytesIO

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('automation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class KITPortalAutomation:
    """Main automation class for KIT portal data extraction"""
    
    def __init__(self, config_path: str = "config.json"):
        """Initialize automation with configuration"""
        self.config = self.load_config(config_path)
        self.driver = None
        self.reader = easyocr.Reader(['en'], gpu=False)
        self.base_url = "https://portal.kitcbe.com/index.php/Login"
        self.password = self.config['portal']['password']
        self.output_dir = Path(self.config['output']['directory'])
        self.output_dir.mkdir(exist_ok=True)
        self.photos_dir = self.output_dir / "photos"
        self.photos_dir.mkdir(exist_ok=True)
        
    def load_config(self, config_path: str) -> Dict:
        """Load configuration from JSON file"""
        try:
            with open(config_path, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            logger.error(f"Config file not found: {config_path}")
            raise
    
    def setup_driver(self):
        """Setup Selenium WebDriver with Chrome"""
        options = Options()
        # options.add_argument('--headless')  # Uncomment for headless mode
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--start-maximized')
        options.add_argument('--disable-gpu')
        
        # Prevent detection
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        self.driver = webdriver.Chrome(options=options)
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        self.driver.implicitly_wait(10)
        logger.info("WebDriver initialized successfully")
    
    def preprocess_captcha_image(self, input_path: str) -> str:
        """Preprocess CAPTCHA image for better OCR - handles overlapping text"""
        try:
            from PIL import Image, ImageEnhance, ImageFilter
            import cv2
            import numpy as np
            
            # Load image
            img = Image.open(input_path)
            
            # Convert to RGB
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Resize for better OCR (2x larger)
            width, height = img.size
            img = img.resize((width * 2, height * 2), Image.Resampling.LANCZOS)
            
            # Enhance contrast aggressively
            enhancer = ImageEnhance.Contrast(img)
            img = enhancer.enhance(3.0)
            
            # Convert to grayscale
            img = img.convert('L')
            
            # Convert to numpy for OpenCV processing
            img_array = np.array(img)
            
            # Apply bilateral filter to reduce noise while keeping edges
            img_array = cv2.bilateralFilter(img_array, 9, 75, 75)
            
            # Adaptive threshold - handles varying lighting
            img_array = cv2.adaptiveThreshold(
                img_array, 255,
                cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY, 11, 2
            )
            
            # Invert if background is dark
            if np.mean(img_array) < 127:
                img_array = cv2.bitwise_not(img_array)
            
            # Morphological operations to clean up
            kernel = np.ones((2, 2), np.uint8)
            img_array = cv2.morphologyEx(img_array, cv2.MORPH_CLOSE, kernel)
            
            # Save processed image
            processed_path = 'captcha_processed.png'
            cv2.imwrite(processed_path, img_array)
            
            return processed_path
            
        except Exception as e:
            logger.warning(f"Image preprocessing failed: {e}, using original")
            return input_path
    
    def solve_captcha(self, max_retries: int = 3) -> Optional[str]:
        """Solve CAPTCHA using OCR with preprocessing"""
        for attempt in range(max_retries):
            try:
                time.sleep(1)
                
                # Find CAPTCHA image
                captcha_img = self.driver.find_element(By.XPATH, 
                    "//img[contains(@src, 'captcha_images')]")
                
                if not captcha_img:
                    logger.warning(f"CAPTCHA image not found on attempt {attempt + 1}")
                    continue
                
                # Take screenshot
                captcha_filename = 'captcha_temp.png'
                captcha_img.screenshot(captcha_filename)
                logger.info(f"CAPTCHA screenshot saved: {captcha_filename}")
                
                # Preprocess image for better OCR
                processed_path = self.preprocess_captcha_image(captcha_filename)
                
                # Try OCR with multiple configurations
                results = []
                
                # Method 1: Default EasyOCR
                result1 = self.reader.readtext(processed_path, detail=0)
                if result1:
                    text1 = ''.join(result1).strip()
                    text1 = re.sub(r'[^A-Za-z0-9]', '', text1)
                    if text1:
                        results.append(text1)
                
                # Method 2: With allowlist (only letters and numbers)
                result2 = self.reader.readtext(
                    processed_path, 
                    detail=0,
                    allowlist='ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789'
                )
                if result2:
                    text2 = ''.join(result2).strip()
                    text2 = re.sub(r'[^A-Za-z0-9]', '', text2)
                    if text2:
                        results.append(text2)
                
                # Method 3: Try original image too
                result3 = self.reader.readtext(captcha_filename, detail=0)
                if result3:
                    text3 = ''.join(result3).strip()
                    text3 = re.sub(r'[^A-Za-z0-9]', '', text3)
                    if text3:
                        results.append(text3)
                
                # Choose best result (prefer length 4-5)
                valid_results = [r for r in results if 4 <= len(r) <= 6]
                
                if valid_results:
                    # Pick most common result if multiple
                    from collections import Counter
                    counter = Counter(valid_results)
                    captcha_text = counter.most_common(1)[0][0]
                    
                    logger.info(f"CAPTCHA solved: {captcha_text} (attempt {attempt + 1})")
                    logger.debug(f"All results: {results}")
                    return captcha_text
                elif results:
                    # Use longest result as fallback
                    captcha_text = max(results, key=len)
                    if len(captcha_text) >= 3:
                        logger.info(f"CAPTCHA (fallback): {captcha_text}")
                        return captcha_text
                
                logger.warning(f"No valid CAPTCHA result on attempt {attempt + 1}")
                logger.debug(f"Results: {results}")
                
            except Exception as e:
                logger.warning(f"CAPTCHA solve attempt {attempt + 1} failed: {e}")
                time.sleep(1)
        
        logger.error("Failed to solve CAPTCHA after all retries")
        return None
    
    def click_login_button(self) -> bool:
        """
        Try multiple methods to click the login button
        FIXED: Handles element not interactable error
        """
        try:
            # Wait a bit for any animations/validations
            time.sleep(1)
            
            # Method 1: Wait for button to be clickable
            try:
                logger.info("Attempting Method 1: Wait for clickable...")
                login_btn = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Login')]"))
                )
                
                # Scroll into view
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", login_btn)
                time.sleep(0.5)
                
                login_btn.click()
                logger.info("✓ Login button clicked (Method 1)")
                return True
                
            except (ElementNotInteractableException, ElementClickInterceptedException) as e:
                logger.warning(f"Method 1 failed: {e}")
            
            # Method 2: JavaScript click
            try:
                logger.info("Attempting Method 2: JavaScript click...")
                login_btn = self.driver.find_element(By.XPATH, "//button[contains(text(), 'Login')]")
                self.driver.execute_script("arguments[0].click();", login_btn)
                logger.info("✓ Login button clicked (Method 2 - JS)")
                return True
                
            except Exception as e:
                logger.warning(f"Method 2 failed: {e}")
            
            # Method 3: ActionChains
            try:
                logger.info("Attempting Method 3: ActionChains...")
                login_btn = self.driver.find_element(By.XPATH, "//button[contains(text(), 'Login')]")
                actions = ActionChains(self.driver)
                actions.move_to_element(login_btn).click().perform()
                logger.info("✓ Login button clicked (Method 3 - Actions)")
                return True
                
            except Exception as e:
                logger.warning(f"Method 3 failed: {e}")
            
            # Method 4: Try by class name
            try:
                logger.info("Attempting Method 4: By class name...")
                login_btn = self.driver.find_element(By.CSS_SELECTOR, "button.login100-form-btn")
                self.driver.execute_script("arguments[0].click();", login_btn)
                logger.info("✓ Login button clicked (Method 4 - Class)")
                return True
                
            except Exception as e:
                logger.warning(f"Method 4 failed: {e}")
            
            # Method 5: Submit form directly
            try:
                logger.info("Attempting Method 5: Submit form...")
                form = self.driver.find_element(By.TAG_NAME, "form")
                form.submit()
                logger.info("✓ Form submitted (Method 5)")
                return True
                
            except Exception as e:
                logger.warning(f"Method 5 failed: {e}")
            
            logger.error("All click methods failed!")
            return False
            
        except Exception as e:
            logger.error(f"Error in click_login_button: {e}")
            return False
    
    def login(self, roll_number: str) -> bool:
        """Login to portal - FIXED with improved click handling"""
        try:
            self.driver.get(self.base_url)
            time.sleep(3)
            
            # Wait for login form to be present
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.ID, "username"))
            )
            
            # Enter username
            logger.info("Finding username field...")
            roll_input = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, "username"))
            )
            
            roll_input.clear()
            time.sleep(0.5)
            roll_input.send_keys(roll_number)
            logger.info(f"✓ Roll number entered: {roll_number}")
            
            # Enter password
            logger.info("Finding password field...")
            password_input = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, "password1"))
            )
            
            password_input.clear()
            time.sleep(0.5)
            password_input.send_keys(self.password)
            logger.info("✓ Password entered")
            
            # Solve CAPTCHA
            captcha_text = self.solve_captcha()
            if not captcha_text:
                logger.error("CAPTCHA solving failed")
                return False
            
            # Enter CAPTCHA
            logger.info("Finding CAPTCHA field...")
            captcha_input = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, "captcha"))
            )
            
            captcha_input.clear()
            time.sleep(0.5)
            captcha_input.send_keys(captcha_text)
            logger.info(f"✓ CAPTCHA entered: {captcha_text}")
            
            # Extra wait for form validation
            time.sleep(1)
            
            # Click login button with improved method
            if not self.click_login_button():
                logger.error("Failed to click login button")
                self.driver.save_screenshot(f"login_click_failed_{roll_number}.png")
                return False
            
            # Wait for page to load
            time.sleep(5)
            
            # Check if login successful
            current_url = self.driver.current_url
            page_source = self.driver.page_source
            
            if ("Results" in current_url or "PROVISIONAL RESULTS" in page_source or 
                "RESULT" in page_source or "Results" in page_source):
                logger.info(f"✓ Login successful for {roll_number}")
                return True
            else:
                logger.error(f"✗ Login failed for {roll_number}")
                logger.debug(f"Current URL: {current_url}")
                
                # Check for error messages
                try:
                    error_elem = self.driver.find_element(By.XPATH, 
                        "//*[contains(text(), 'Invalid') or contains(text(), 'incorrect') or contains(text(), 'wrong')]")
                    logger.error(f"Error message: {error_elem.text}")
                except:
                    pass
                    
                # Save screenshot for debugging
                self.driver.save_screenshot(f"login_failed_{roll_number}.png")
                logger.error(f"Screenshot saved: login_failed_{roll_number}.png")
                
                return False
                
        except Exception as e:
            logger.error(f"Login error for {roll_number}: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def extract_marksheet_data(self) -> Dict:
        """Extract data from marksheet/results page"""
        data = {}
        try:
            time.sleep(2)
            
            # Extract basic info
            try:
                data['name'] = self.driver.find_element(By.XPATH, 
                    "//td[contains(text(), 'Name')]/following-sibling::td").text.strip()
            except:
                data['name'] = ''
            
            try:
                data['register_number'] = self.driver.find_element(By.XPATH, 
                    "//td[contains(text(), 'Register Number')]/following-sibling::td").text.strip()
            except:
                data['register_number'] = ''
            
            try:
                data['regulation'] = self.driver.find_element(By.XPATH, 
                    "//td[contains(text(), 'Regulation')]/following-sibling::td").text.strip()
            except:
                data['regulation'] = ''
            
            try:
                data['gender'] = self.driver.find_element(By.XPATH, 
                    "//td[contains(text(), 'Gender')]/following-sibling::td").text.strip()
            except:
                data['gender'] = ''
            
            try:
                data['dob'] = self.driver.find_element(By.XPATH, 
                    "//td[contains(text(), 'Date of Birth')]/following-sibling::td").text.strip()
            except:
                data['dob'] = ''
            
            try:
                data['branch'] = self.driver.find_element(By.XPATH, 
                    "//td[contains(text(), 'Branch')]/following-sibling::td").text.strip()
            except:
                data['branch'] = ''
            
            # Extract marks table
            courses = []
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, 
                        "//table[.//th[contains(text(), 'COURSE NAME') or contains(text(), 'Course Name')]]"))
                )
                
                marks_table = self.driver.find_element(By.XPATH, 
                    "//table[.//th[contains(text(), 'COURSE NAME') or contains(text(), 'Course Name')]]")
                rows = marks_table.find_elements(By.TAG_NAME, "tr")[1:]
                
                for row in rows:
                    cols = row.find_elements(By.TAG_NAME, "td")
                    if len(cols) >= 4:
                        course_data = {
                            'semester': cols[0].text.strip() if len(cols) > 0 else '',
                            'course_code': cols[1].text.strip() if len(cols) > 1 else '',
                            'course_name': cols[2].text.strip() if len(cols) > 2 else '',
                            'grade': cols[3].text.strip() if len(cols) > 3 else '',
                            'gp': cols[4].text.strip() if len(cols) > 4 else '',
                            'result': cols[5].text.strip() if len(cols) > 5 else ''
                        }
                        if course_data['course_name']:
                            courses.append(course_data)
                
                logger.info(f"Extracted {len(courses)} courses")
            except Exception as e:
                logger.warning(f"Could not extract marks table: {e}")
            
            data['courses'] = courses
            logger.info(f"✓ Extracted marksheet data for {data.get('register_number', 'Unknown')}")
            
        except Exception as e:
            logger.error(f"Error extracting marksheet data: {e}")
        
        return data
    
    def navigate_to_profile(self) -> bool:
        """Navigate to profile details page"""
        try:
            time.sleep(2)
            
            # Click on profile area (top right)
            profile_clicked = False
            
            # Method 1: Look for student name/profile text in header
            try:
                profile_elem = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, 
                        "//*[contains(@class, 'profile') or contains(text(), 'STUDENTS')]"))
                )
                self.driver.execute_script("arguments[0].scrollIntoView(true);", profile_elem)
                time.sleep(0.5)
                
                # Try normal click first
                try:
                    profile_elem.click()
                    profile_clicked = True
                    logger.info("Clicked profile area (method 1)")
                except:
                    # Fallback to JS click
                    self.driver.execute_script("arguments[0].click();", profile_elem)
                    profile_clicked = True
                    logger.info("Clicked profile area (method 1 - JS)")
            except:
                pass
            
            # Method 2: Look for profile image
            if not profile_clicked:
                try:
                    profile_img = self.driver.find_element(By.XPATH, 
                        "//img[contains(@alt, 'profile') or contains(@class, 'profile')]")
                    self.driver.execute_script("arguments[0].click();", profile_img)
                    profile_clicked = True
                    logger.info("Clicked profile area (method 2)")
                except:
                    pass
            
            if not profile_clicked:
                logger.error("Could not click profile area")
                self.driver.save_screenshot("profile_navigation_failed.png")
                return False
            
            time.sleep(2)
            
            # Click "Profile Details" from dropdown
            try:
                profile_link = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, 
                        "//a[contains(text(), 'Profile Details') or contains(text(), 'Profile')]"))
                )
                
                # Try normal click first
                try:
                    profile_link.click()
                    logger.info("Clicked 'Profile Details' link")
                except:
                    # Fallback to JS click
                    self.driver.execute_script("arguments[0].click();", profile_link)
                    logger.info("Clicked 'Profile Details' link (JS)")
                    
            except Exception as e:
                logger.error(f"Could not click Profile Details link: {e}")
                self.driver.save_screenshot("profile_dropdown_failed.png")
                return False
            
            time.sleep(3)
            
            # Verify we're on profile page
            current_url = self.driver.current_url
            page_source = self.driver.page_source
            
            if ("Usersprofile" in current_url or "Edit User" in page_source or 
                "First Name" in page_source):
                logger.info("✓ Navigated to profile page")
                return True
            else:
                logger.error(f"✗ Profile page not loaded. URL: {current_url}")
                self.driver.save_screenshot("wrong_profile_page.png")
                return False
            
        except Exception as e:
            logger.error(f"Error navigating to profile: {e}")
            return False
    
    def extract_profile_data(self, roll_number: str) -> Dict:
        """Extract all data from profile page"""
        data = {}
        try:
            time.sleep(2)
            
            # Download profile photo
            try:
                photo_elem = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, 
                        "//img[contains(@src, 'upload') or contains(@class, 'profile')]"))
                )
                
                photo_url = photo_elem.get_attribute('src')
                
                # Handle relative URLs
                if not photo_url.startswith('http'):
                    base = 'https://portal.kitcbe.com'
                    if photo_url.startswith('/'):
                        photo_url = base + photo_url
                    else:
                        photo_url = base + '/' + photo_url
                
                logger.info(f"Downloading photo from: {photo_url}")
                
                response = requests.get(photo_url, timeout=10)
                if response.status_code == 200:
                    photo_path = self.photos_dir / f"{roll_number}.jpg"
                    with open(photo_path, 'wb') as f:
                        f.write(response.content)
                    
                    data['photo_path'] = str(photo_path)
                    logger.info(f"✓ Downloaded photo for {roll_number}")
                else:
                    logger.warning(f"Photo download failed: HTTP {response.status_code}")
                    data['photo_path'] = None
                
            except Exception as e:
                logger.warning(f"Could not download photo: {e}")
                data['photo_path'] = None
            
            # Extract form fields
            form_fields = {
                'first_name': "First Name",
                'last_name': "Last Name",
                'blood_group': "Blood Group",
                'mobile': "Mobile Number",
                'email': "Email",
                'marital_status': "Marital Status",
                'alternative_mobile': "Alternative Mobile Number",
                'alternative_email': "Alternative Email",
                'pan': "PAN No",
                'aadhaar': "Aadhaar No",
                'geographic_classification': "Geographic Classification",
                'community': "Community",
                'caste': "Caste",
                'religion': "Religion",
                'mother_tongue': "Mother Tongue",
                'nationality': "Nationality",
                'nativity': "Nativity",
                'profile_status': "Profile Status",
            }
            
            for key, label in form_fields.items():
                try:
                    value = None
                    
                    # Try input field
                    try:
                        elem = self.driver.find_element(By.XPATH, 
                            f"//input[preceding-sibling::label[contains(text(), '{label}')] or "
                            f"following-sibling::label[contains(text(), '{label}')]]")
                        value = elem.get_attribute('value')
                    except:
                        pass
                    
                    # Try select/dropdown
                    if not value:
                        try:
                            elem = self.driver.find_element(By.XPATH, 
                                f"//select[preceding-sibling::label[contains(text(), '{label}')] or "
                                f"following-sibling::label[contains(text(), '{label}')]]")
                            selected = elem.find_element(By.XPATH, ".//option[@selected]")
                            value = selected.text
                        except:
                            pass
                    
                    # Try finding in same row/div
                    if not value:
                        try:
                            label_elem = self.driver.find_element(By.XPATH, 
                                f"//label[contains(text(), '{label}')]")
                            parent = label_elem.find_element(By.XPATH, "./..")
                            
                            try:
                                input_elem = parent.find_element(By.TAG_NAME, "input")
                                value = input_elem.get_attribute('value')
                            except:
                                pass
                            
                            if not value:
                                try:
                                    select_elem = parent.find_element(By.TAG_NAME, "select")
                                    selected = select_elem.find_element(By.XPATH, ".//option[@selected]")
                                    value = selected.text
                                except:
                                    pass
                        except:
                            pass
                    
                    # Skip if empty or "Select"
                    if value and value.strip() and value.strip().lower() not in ['select', '']:
                        data[key] = value.strip()
                    else:
                        data[key] = ''
                    
                except Exception as e:
                    logger.debug(f"Field {key} not found: {e}")
                    data[key] = ''
            
            # Handle Gender (radio button)
            try:
                gender_elem = self.driver.find_element(By.XPATH, 
                    "//input[@type='radio' and @checked and (@value='MALE' or @value='FEMALE' or @value='TRANSGENDER')]")
                data['gender_profile'] = gender_elem.get_attribute('value')
            except:
                data['gender_profile'] = ''
            
            # Handle Date of Birth
            try:
                dob_elem = self.driver.find_element(By.XPATH, 
                    "//input[contains(@placeholder, 'dd-mm-yyyy') or @type='date']")
                data['dob_profile'] = dob_elem.get_attribute('value')
            except:
                data['dob_profile'] = ''
            
            logger.info(f"✓ Extracted profile data for {roll_number}")
            filled_count = sum(1 for v in data.values() if v and v != '')
            logger.info(f"Profile fields filled: {filled_count}/{len(data)}")
            
        except Exception as e:
            logger.error(f"Error extracting profile data: {e}")
        
        return data
    
    def logout(self):
        """Logout from portal"""
        try:
            time.sleep(1)
            
            # Click profile area
            try:
                profile_elem = self.driver.find_element(By.XPATH, 
                    "//*[contains(@class, 'profile') or contains(text(), 'STUDENTS')]")
                self.driver.execute_script("arguments[0].click();", profile_elem)
                time.sleep(1)
            except:
                try:
                    imgs = self.driver.find_elements(By.TAG_NAME, "img")
                    for img in imgs:
                        if img.size['height'] < 100:
                            self.driver.execute_script("arguments[0].click();", img)
                            time.sleep(1)
                            break
                except:
                    pass
            
            # Click logout
            try:
                logout_link = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, 
                        "//a[contains(text(), 'Logout') or contains(text(), 'logout')]"))
                )
                self.driver.execute_script("arguments[0].click();", logout_link)
                time.sleep(2)
                logger.info("✓ Logged out successfully")
            except:
                # Force logout by navigating to logout URL
                try:
                    self.driver.get("https://portal.kitcbe.com/index.php/Login/logout")
                    time.sleep(2)
                    logger.info("✓ Logged out (forced)")
                except:
                    pass
            
        except Exception as e:
            logger.warning(f"Logout error: {e}")
    
    def process_student(self, roll_number: str) -> Dict:
        """Process single student"""
        logger.info(f"\n{'='*60}")
        logger.info(f"Processing: {roll_number}")
        logger.info(f"{'='*60}")
        
        student_data = {'roll_number': roll_number}
        
        try:
            # Login
            if not self.login(roll_number):
                student_data['status'] = 'Login Failed'
                return student_data
            
            # Extract marksheet data
            marksheet_data = self.extract_marksheet_data()
            student_data.update(marksheet_data)
            
            # Navigate to profile
            if self.navigate_to_profile():
                profile_data = self.extract_profile_data(roll_number)
                student_data.update(profile_data)
                student_data['status'] = 'Success'
            else:
                student_data['status'] = 'Profile Navigation Failed'
            
            # Logout
            self.logout()
            
            logger.info(f"✓ Successfully processed {roll_number}")
            
        except Exception as e:
            logger.error(f"✗ Error processing {roll_number}: {e}")
            student_data['status'] = f'Error: {str(e)}'
        
        return student_data
    
    def generate_roll_numbers(self, department_config: Dict) -> List[str]:
        """Generate list of roll numbers"""
        prefix = department_config['prefix']
        start = department_config['start']
        end = department_config['end']
        
        roll_numbers = [f"{prefix}{i:03d}" for i in range(start, end + 1)]
        logger.info(f"Generated {len(roll_numbers)} roll numbers")
        return roll_numbers
    
    def save_to_excel(self, all_data: List[Dict], output_file: str):
        """Save all data to Excel with embedded photos"""
        logger.info("Saving data to Excel...")
        
        rows = []
        for student in all_data:
            row = {
                'Roll Number': student.get('roll_number', ''),
                'Status': student.get('status', ''),
                'Name': student.get('name', ''),
                'Register Number': student.get('register_number', ''),
                'First Name': student.get('first_name', ''),
                'Last Name': student.get('last_name', ''),
                'Gender': student.get('gender', ''),
                'DOB': student.get('dob', ''),
                'Blood Group': student.get('blood_group', ''),
                'Branch': student.get('branch', ''),
                'Regulation': student.get('regulation', ''),
                'Mobile': student.get('mobile', ''),
                'Email': student.get('email', ''),
                'Alternative Mobile': student.get('alternative_mobile', ''),
                'Alternative Email': student.get('alternative_email', ''),
                'PAN': student.get('pan', ''),
                'Aadhaar': student.get('aadhaar', ''),
                'Marital Status': student.get('marital_status', ''),
                'Community': student.get('community', ''),
                'Caste': student.get('caste', ''),
                'Religion': student.get('religion', ''),
                'Mother Tongue': student.get('mother_tongue', ''),
                'Nationality': student.get('nationality', ''),
                'Nativity': student.get('nativity', ''),
                'Geographic Classification': student.get('geographic_classification', ''),
                'Profile Status': student.get('profile_status', ''),
                'Photo Path': student.get('photo_path', ''),
            }
            
            # Add course details
            courses = student.get('courses', [])
            for idx, course in enumerate(courses, 1):
                row[f'Course {idx} Semester'] = course.get('semester', '')
                row[f'Course {idx} Code'] = course.get('course_code', '')
                row[f'Course {idx} Name'] = course.get('course_name', '')
                row[f'Course {idx} Grade'] = course.get('grade', '')
                row[f'Course {idx} GP'] = course.get('gp', '')
                row[f'Course {idx} Result'] = course.get('result', '')
            
            rows.append(row)
        
        df = pd.DataFrame(rows)
        output_path = self.output_dir / output_file
        
        df.to_excel(output_path, index=False, engine='openpyxl')
        logger.info(f"✓ Data saved to Excel")
        
        # Embed photos
        try:
            wb = load_workbook(output_path)
            ws = wb.active
            
            ws.insert_cols(1)
            ws.cell(1, 1, 'Photo')
            ws.column_dimensions['A'].width = 15
            
            for idx, student in enumerate(all_data, start=2):
                photo_path = student.get('photo_path')
                if photo_path and os.path.exists(photo_path):
                    try:
                        img = Image.open(photo_path)
                        img.thumbnail((100, 100))
                        temp_path = f"temp_{idx}.jpg"
                        img.save(temp_path)
                        
                        xl_img = XLImage(temp_path)
                        xl_img.width = 80
                        xl_img.height = 80
                        
                        ws.add_image(xl_img, f'A{idx}')
                        ws.row_dimensions[idx].height = 60
                        
                        os.remove(temp_path)
                    except Exception as e:
                        logger.warning(f"Could not embed photo for row {idx}: {e}")
            
            wb.save(output_path)
            logger.info(f"✓ Photos embedded")
        except Exception as e:
            logger.warning(f"Could not embed photos: {e}")
        
        logger.info(f"✓ Saved to {output_path}")
        success = sum(1 for s in all_data if s.get('status') == 'Success')
        logger.info(f"✓ Successful: {success}/{len(all_data)}")
    
    def run(self, department_key: str):
        """Main execution"""
        logger.info("="*80)
        logger.info(f"KIT PORTAL AUTOMATION STARTED")
        logger.info(f"Department: {department_key}")
        logger.info(f"Time: {datetime.now()}")
        logger.info("="*80)
        
        try:
            self.setup_driver()
            
            dept_config = self.config['departments'][department_key]
            roll_numbers = self.generate_roll_numbers(dept_config)
            
            all_data = []
            success_count = 0
            failed_count = 0
            
            for idx, roll_number in enumerate(roll_numbers, 1):
                logger.info(f"\nProgress: {idx}/{len(roll_numbers)}")
                
                student_data = self.process_student(roll_number)
                all_data.append(student_data)
                
                if student_data.get('status') == 'Success':
                    success_count += 1
                else:
                    failed_count += 1
                
                # Small delay between students
                time.sleep(2)
            
            # Save to Excel
            output_filename = f"{department_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            self.save_to_excel(all_data, output_filename)
            
            # Summary
            logger.info("\n" + "="*80)
            logger.info("AUTOMATION COMPLETED")
            logger.info(f"Total: {len(roll_numbers)} | Success: {success_count} | Failed: {failed_count}")
            logger.info("="*80)
            
        except Exception as e:
            logger.error(f"Critical error: {e}", exc_info=True)
        
        finally:
            if self.driver:
                self.driver.quit()
                logger.info("Browser closed")


if __name__ == "__main__":
    automation = KITPortalAutomation("config.json")
    automation.run("aids")  # Change to "cse", "ece", etc.